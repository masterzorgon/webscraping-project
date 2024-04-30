import openpyxl as xl_lib
from openpyxl.styles import Font
import requests
from bs4 import BeautifulSoup

webpage_url = 'https://www.boxofficemojo.com/year/2024/'
website_response = requests.get(webpage_url)
website_content = website_response.content

soup_obj = BeautifulSoup(website_content, 'html.parser')
page_title = soup_obj.title
table_data = soup_obj.findAll("tr")

workbook = xl_lib.Workbook()
active_sheet = workbook.active
active_sheet.title = "Box Office Report"
header_row = ["No.", "Movie Title", "Release Date", "Total Gross", "Theaters", "Average per theater"]
output_sheet = workbook["Box Office Report"]
output_sheet.append(header_row)

output_sheet.cell(1, 1).font = Font(size=16, bold=True)
output_sheet.cell(1, 2).font = Font(size=16, bold=True)
output_sheet.cell(1, 3).font = Font(size=16, bold=True)
output_sheet.cell(1, 4).font = Font(size=16, bold=True)
output_sheet.cell(1, 5).font = Font(size=16, bold=True)
output_sheet.cell(1, 6).font = Font(size=16, bold=True)

for row in table_data[1:6]:
    table_cells = row.findAll("td")
    ranking = float(table_cells[0].text)
    movie_name = table_cells[1].text
    total_revenue = float(table_cells[7].text.replace(",", "").replace("$", ""))
    theater_count = float(table_cells[6].text.replace(",", ""))
    release_date = table_cells[8].text
    average_revenue = round(total_revenue / theater_count, 2)
    row_data = [ranking, movie_name, release_date, total_revenue, theater_count, average_revenue]
    output_sheet.append(row_data)

output_sheet.column_dimensions["A"].width = 5
output_sheet.column_dimensions["B"].width = 30
output_sheet.column_dimensions["C"].width = 16
output_sheet.column_dimensions["D"].width = 16
output_sheet.column_dimensions["E"].width = 16
output_sheet.column_dimensions["F"].width = 16

for cell in output_sheet["E:E"]:
    cell.number_format = '#,##0'

for cell in output_sheet["D:D"]:
    cell.number_format = u'"$ "#,##0.00'

for cell in output_sheet["F:F"]:
    cell.number_format = u'"$ "#,##0.00'

workbook.save("Movies.xlsx")