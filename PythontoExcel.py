import openpyxl as xl
from openpyxl.styles import Font 

#create a new excel documnt
wb = xl.Workbook()
ws = wb.active
#change th title of the worksheet
ws.title = "First Sheet"
#create a new worksheet
wb.create_sheet(index=1,title="Second Sheet")
#write content to a cell 
ws["A1"] = "Invoice" 
ws["A1"].font = Font(name="Times New Roman", size = 24, italic=False, bold = True)
#alternatively
fontobject = Font(name="Times New Roman", size = 24, italic=False, bold = True)
ws["A1"].font = fontobject
ws["A2"] = "Tires"
ws["A3"] = "Brakes"
ws["A4"] = "Alignment"

ws.merge_cells("A1:B1")

#unmerge cells 
#ws.unmerge_cells("A1:B1")

ws["B2"] = 450
ws["B3"] = 225
ws["B4"] = 150 
ws["A8"] = "Total"
ws["A8"].font = Font(size=16,bold = True)
ws["B8"] = '=SUM(B2:B4)'


ws.column_dimensions["A"].width=25

write_sheet = wb["Second Sheet"]
read_wb = xl.load_workbook("ProduceReport.xlsx")
read_ws = read_wb["ProduceReport"]
#G2-J2
#A1-D1, 41


for row in read_ws.iter_rows():
    ls = [i.value for i in row] #i is every column in that row, going one row at a time, i by itself is a cell object, not the value, so i.value is needed
    write_sheet.append(ls)

last_row = write_sheet.max_row

write_sheet.cell(last_row+2,2).value = "Total"
write_sheet.cell(last_row+2,2).font = Font(size=16, bold=True)
write_sheet.cell(last_row+2,3).value = '=SUM(C2:C'+str(last_row) +")"
write_sheet.cell(last_row+2,4).value = '=SUM(D2:D'+str(last_row) +")"


write_sheet.cell(last_row+4,2).value = "Average"
write_sheet.cell(last_row+4,2).font = Font(size=16, bold=True)
write_sheet.cell(last_row+4,3).value = '=AVERAGE(C2:C'+str(last_row) +")"
write_sheet.cell(last_row+4,4).value = '=AVERAGE(D2:D'+str(last_row) +")"

write_sheet.column_dimensions["A"].width = 16
write_sheet.column_dimensions["B"].width = 16
write_sheet.column_dimensions["C"].width = 16
write_sheet.column_dimensions["D"].width = 16


for cell in write_sheet["C:C"]:
    cell.number_format = '#,##0'

for cell in write_sheet["D:D"]:
    cell.number_format = u'"$ "#,##0.00'


wb.save("PythontoExcel.xlsx")







