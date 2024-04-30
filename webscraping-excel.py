import openpyxl as xl_lib

workbook_obj = xl_lib.load_workbook('example.xlsx')
sheet_names = workbook_obj.sheetnames
print(sheet_names)

active_sheet = workbook_obj["Sheet1"]
cell_a1 = active_sheet["A1"]
print(active_sheet)
print(cell_a1)
print(cell_a1.value)
print(cell_a1.row)
print(cell_a1.column)
print(cell_a1.coordinate)
print(active_sheet.cell(1, 2).value)
print(active_sheet.max_row)
print(active_sheet.max_column)

counter_var = 1
for iteration in range(7):
    print(active_sheet.cell(counter_var, 2).value)
    counter_var += 1

print(xl_lib.utils.get_column_letter(900))
print(xl_lib.utils.column_index_from_string('AHP'))

for row_cells in active_sheet["A1":"C3"]:
    for current_cell in row_cells:
        print(current_cell.coordinate, current_cell.value)

for row_obj in active_sheet.iter_rows(min_row=1, max_row=active_sheet.max_row, max_col=active_sheet.max_column):
    print(row_obj)
    print(row_obj[0].value)
    print(row_obj[1].value)
    print(row_obj[2].value)

